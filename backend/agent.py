"""
SOC1 Type II Report Processing Agent

This module provides functionality to:
1. Extract text and tables from PDF Type II reports
2. Read Excel management review templates
3. Use Google Gemini AI to intelligently map PDF content to Excel fields
4. Return a filled-out Excel management review
"""

from __future__ import annotations

import gc
import json
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import psutil
from google import genai
from google.genai import types
import openpyxl
import pdfplumber
from dotenv import load_dotenv

# xlsxwriter imported locally in fill_template() to avoid startup cost


def log_memory(label: str) -> float:
    """Log current process RSS memory usage. Returns RSS in MB."""
    process = psutil.Process(os.getpid())
    rss_mb = process.memory_info().rss / (1024 * 1024)
    print(f"[MEMORY] {label}: {rss_mb:.1f} MB RSS")
    return rss_mb

# XML namespaces used when parsing XLSX internals
_SSML = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_ROFF = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

load_dotenv()

GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
if not GOOGLE_API_KEY:
    raise ValueError("GOOGLE_API_KEY not found in environment variables. Please set it in .env file.")


@dataclass
class ExtractedPDFContent:
    """Container for extracted PDF content."""

    full_text: str
    pages: list[str]
    tables: list[list[list[str]]]
    metadata: dict[str, Any]


@dataclass
class ExcelTemplate:
    """Container for Excel template structure."""

    filepath: Path
    sheet_names: list[str]
    headers: dict[str, list[str]]  # sheet_name -> list of column headers
    header_to_col: dict[str, dict[str, int]]  # sheet_name -> {header: col_index}
    header_row: dict[str, int]  # sheet_name -> header row number
    structure: dict[str, list[dict[str, Any]]]  # sheet_name -> list of row dicts
    sheet_type: dict[str, str] = field(default_factory=dict)  # "form" or "table"
    form_fields: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    # Layout info captured from ZIP for xlsxwriter output
    column_widths: dict[str, dict[int, float]] = field(default_factory=dict)
    merged_ranges: dict[str, list[str]] = field(default_factory=dict)
    existing_cells: dict[str, dict[tuple[int, int], Any]] = field(default_factory=dict)
    max_row: dict[str, int] = field(default_factory=dict)
    max_col: dict[str, int] = field(default_factory=dict)
    # Formatting info (fonts, colors, borders, etc.)
    row_heights: dict[str, dict[int, float]] = field(default_factory=dict)
    cell_formats: dict[str, dict[tuple[int, int], dict[str, Any]]] = field(default_factory=dict)
    style_definitions: dict[str, Any] = field(default_factory=dict)  # Parsed from styles.xml


class PDFExtractor:
    """Extracts text and tables from PDF documents using pdfplumber."""

    @staticmethod
    def extract(pdf_path: Path) -> ExtractedPDFContent:
        """
        Extract all text and tables from a PDF file.

        Args:
            pdf_path: Path to the PDF file

        Returns:
            ExtractedPDFContent with full text, per-page text, tables, and metadata
        """
        # MEMORY FIX: Build full_text directly instead of storing all pages
        text_parts: list[str] = []
        tables: list[list[list[str]]] = []
        metadata: dict[str, Any] = {}

        with pdfplumber.open(pdf_path) as pdf:
            metadata = {
                "num_pages": len(pdf.pages),
                "metadata": pdf.metadata or {},
            }

            # PERFORMANCE: Expanded to 150 pages for better coverage (was 50)
            # Most SOC1 reports are 100-200 pages - we want full context
            max_pages = min(150, len(pdf.pages))

            for i, page in enumerate(pdf.pages[:max_pages]):
                # Extract text from page
                page_text = page.extract_text() or ""
                text_parts.append(page_text)

                # PERFORMANCE: Expanded to 40 tables (was 10) for comprehensive extraction
                if len(tables) < 40:
                    page_tables = page.extract_tables() or []
                    for table in page_tables:
                        if len(tables) >= 40:
                            break
                        # Clean up table data
                        cleaned_table = [
                            [str(cell) if cell is not None else "" for cell in row]
                            for row in table
                        ]
                        tables.append(cleaned_table)

        full_text = "\n\n--- Page Break ---\n\n".join(text_parts)

        # MEMORY FIX: Delete text_parts immediately to avoid duplicate storage
        del text_parts
        gc.collect()

        # MEMORY FIX: Don't store individual pages, just return empty list
        # This saves significant memory for large PDFs
        return ExtractedPDFContent(
            full_text=full_text,
            pages=[],  # Empty to save memory
            tables=tables,
            metadata=metadata,
        )


class ExcelHandler:
    """Handles reading and writing Excel files.

    Read phase:  openpyxl read_only  (streaming, ~constant memory)
    Write phase: xlsxwriter          (forward-only, ~constant memory)

    The XLSX ZIP is parsed directly for column widths and merged cells,
    avoiding the need to ever open the workbook in writable mode with
    openpyxl (which can use 50x the file size in RAM).
    """

    # ── helpers ──────────────────────────────────────────────────────

    @staticmethod
    def _ref_to_rowcol(ref: str) -> tuple[int, int]:
        """Convert Excel cell reference like 'A1' to 1-based (row, col)."""
        col_str = ""
        row_str = ""
        for c in ref:
            if c.isalpha():
                col_str += c
            else:
                row_str += c
        col = 0
        for c in col_str.upper():
            col = col * 26 + (ord(c) - ord('A') + 1)
        return int(row_str), col

    @staticmethod
    def _parse_range(range_ref: str) -> tuple[int, int, int, int]:
        """Parse 'A1:B3' into (first_row, first_col, last_row, last_col) 1-based."""
        parts = range_ref.split(":")
        r1, c1 = ExcelHandler._ref_to_rowcol(parts[0])
        r2, c2 = ExcelHandler._ref_to_rowcol(parts[1]) if len(parts) > 1 else (r1, c1)
        return r1, c1, r2, c2

    @staticmethod
    def _match_sheet_name(
        ai_name: str, template_names: list[str]
    ) -> str | None:
        """Fuzzy-match an AI-returned sheet name to a template sheet name."""
        # Exact match
        if ai_name in template_names:
            return ai_name
        ai_lower = ai_name.lower()
        for tpl_name in template_names:
            tpl_lower = tpl_name.lower()
            # Substring match (either direction)
            if ai_lower in tpl_lower or tpl_lower in ai_lower:
                return tpl_name
            # Keyword match
            for kw in ("management review", "user entity", "cuec", "comp user"):
                if kw in ai_lower and kw in tpl_lower:
                    return tpl_name
        return None

    # ── ZIP layout parser (streaming, low memory) ────────────────────

    @staticmethod
    def _parse_styles_from_zip(zf: zipfile.ZipFile) -> dict[str, Any]:
        """
        Parse xl/styles.xml to extract fonts, fills, borders, number formats,
        and cellXfs (cell format definitions).
        Returns a dict with parsed style information.
        """
        styles_data = {
            "fonts": [],
            "fills": [],
            "borders": [],
            "numFmts": {},
            "cellXfs": [],
        }

        try:
            if "xl/styles.xml" not in zf.namelist():
                return styles_data

            styles_bytes = zf.read("xl/styles.xml")
            styles_root = ET.fromstring(styles_bytes)
            del styles_bytes

            # Parse fonts
            fonts_el = styles_root.find(f"{_SSML}fonts")
            if fonts_el is not None:
                for font_el in fonts_el.findall(f"{_SSML}font"):
                    font = {}
                    name_el = font_el.find(f"{_SSML}name")
                    if name_el is not None:
                        font["name"] = name_el.get("val", "Calibri")
                    sz_el = font_el.find(f"{_SSML}sz")
                    if sz_el is not None:
                        font["size"] = float(sz_el.get("val", "11"))
                    font["bold"] = font_el.find(f"{_SSML}b") is not None
                    font["italic"] = font_el.find(f"{_SSML}i") is not None
                    color_el = font_el.find(f"{_SSML}color")
                    if color_el is not None:
                        rgb = color_el.get("rgb")
                        if rgb and len(rgb) >= 6:
                            font["color"] = rgb[-6:]  # Get last 6 chars (RRGGBB)
                    styles_data["fonts"].append(font)

            # Parse fills
            fills_el = styles_root.find(f"{_SSML}fills")
            if fills_el is not None:
                for fill_el in fills_el.findall(f"{_SSML}fill"):
                    fill = {}
                    pat_fill = fill_el.find(f"{_SSML}patternFill")
                    if pat_fill is not None:
                        fill["patternType"] = pat_fill.get("patternType")
                        fg_color = pat_fill.find(f"{_SSML}fgColor")
                        if fg_color is not None:
                            rgb = fg_color.get("rgb")
                            if rgb and len(rgb) >= 6:
                                fill["fgColor"] = rgb[-6:]
                    styles_data["fills"].append(fill)

            # Parse borders
            borders_el = styles_root.find(f"{_SSML}borders")
            if borders_el is not None:
                for border_el in borders_el.findall(f"{_SSML}border"):
                    border = {}
                    for side in ["left", "right", "top", "bottom"]:
                        side_el = border_el.find(f"{_SSML}{side}")
                        if side_el is not None:
                            style = side_el.get("style")
                            if style:
                                border[side] = style
                    styles_data["borders"].append(border)

            # Parse number formats
            numFmts_el = styles_root.find(f"{_SSML}numFmts")
            if numFmts_el is not None:
                for numFmt_el in numFmts_el.findall(f"{_SSML}numFmt"):
                    fmt_id = numFmt_el.get("numFmtId")
                    fmt_code = numFmt_el.get("formatCode")
                    if fmt_id and fmt_code:
                        styles_data["numFmts"][int(fmt_id)] = fmt_code

            # Parse cellXfs (cell format definitions)
            cellXfs_el = styles_root.find(f"{_SSML}cellXfs")
            if cellXfs_el is not None:
                for xf_el in cellXfs_el.findall(f"{_SSML}xf"):
                    xf = {}
                    font_id = xf_el.get("fontId")
                    fill_id = xf_el.get("fillId")
                    border_id = xf_el.get("borderId")
                    num_fmt_id = xf_el.get("numFmtId")

                    if font_id is not None:
                        xf["fontId"] = int(font_id)
                    if fill_id is not None:
                        xf["fillId"] = int(fill_id)
                    if border_id is not None:
                        xf["borderId"] = int(border_id)
                    if num_fmt_id is not None:
                        xf["numFmtId"] = int(num_fmt_id)

                    # Parse alignment
                    align_el = xf_el.find(f"{_SSML}alignment")
                    if align_el is not None:
                        alignment = {}
                        if align_el.get("horizontal"):
                            alignment["horizontal"] = align_el.get("horizontal")
                        if align_el.get("vertical"):
                            alignment["vertical"] = align_el.get("vertical")
                        if align_el.get("wrapText") == "1":
                            alignment["wrapText"] = True
                        if alignment:
                            xf["alignment"] = alignment

                    styles_data["cellXfs"].append(xf)

            del styles_root
        except Exception as e:
            print(f"Warning: Could not parse styles.xml: {e}")

        return styles_data

    @staticmethod
    def _parse_layout_from_zip(
        excel_path: Path,
        target_sheet_names: list[str],
    ) -> tuple[dict[str, dict[int, float]], dict[str, list[str]], dict[str, dict[int, float]], dict[str, dict[tuple[int, int], dict[str, Any]]], dict[str, Any]]:
        """
        Parse column widths, merged cells, row heights, and cell formatting
        directly from the XLSX ZIP using streaming XML (iterparse).
        Never loads the full sheet DOM into memory.

        Returns:
            - column_widths: sheet -> {col_idx: width}
            - merged_ranges: sheet -> [range_refs]
            - row_heights: sheet -> {row_idx: height}
            - cell_formats: sheet -> {(row, col): {styleId, ...}}
            - style_definitions: parsed styles from styles.xml
        """
        column_widths: dict[str, dict[int, float]] = {}
        merged_ranges: dict[str, list[str]] = {}
        row_heights: dict[str, dict[int, float]] = {}
        cell_formats: dict[str, dict[tuple[int, int], dict[str, Any]]] = {}
        style_definitions: dict[str, Any] = {}

        try:
            with zipfile.ZipFile(excel_path, "r") as zf:
                # Parse styles first
                style_definitions = ExcelHandler._parse_styles_from_zip(zf)

                # Map sheet names ➜ XML file paths inside the ZIP
                wb_bytes = zf.read("xl/workbook.xml")
                wb_root = ET.fromstring(wb_bytes)
                del wb_bytes
                sheets_el = wb_root.find(f"{_SSML}sheets")
                if sheets_el is None:
                    return column_widths, merged_ranges, row_heights, cell_formats, style_definitions

                sheet_rid: dict[str, str] = {}
                for s in sheets_el.findall(f"{_SSML}sheet"):
                    name = s.get("name")
                    rid = s.get(f"{_ROFF}id")
                    if name in target_sheet_names and rid:
                        sheet_rid[name] = rid
                del wb_root

                rels_bytes = zf.read("xl/_rels/workbook.xml.rels")
                rels_root = ET.fromstring(rels_bytes)
                del rels_bytes
                rid_to_path: dict[str, str] = {}
                for rel in rels_root.findall(f"{_REL}Relationship"):
                    rid_to_path[rel.get("Id")] = rel.get("Target")
                del rels_root

                # Stream-parse each target sheet's XML
                for sheet_name, rid in sheet_rid.items():
                    target = rid_to_path.get(rid)
                    if not target:
                        continue
                    xml_path = f"xl/{target}" if not target.startswith("/") else target.lstrip("/")
                    if xml_path not in zf.namelist():
                        continue

                    widths: dict[int, float] = {}
                    merges: list[str] = []
                    heights: dict[int, float] = {}
                    formats: dict[tuple[int, int], dict[str, Any]] = {}

                    # MEMORY OPTIMIZATION: Only capture styles for first 60 rows
                    # Actual data is ~36 rows, using 60 for safety buffer
                    MAX_STYLE_ROWS = 60

                    with zf.open(xml_path) as f:
                        for event, elem in ET.iterparse(f, events=("end",)):
                            tag = elem.tag
                            if tag == f"{_SSML}col":
                                min_c = int(elem.get("min", "1"))
                                max_c = int(elem.get("max", str(min_c)))
                                w = float(elem.get("width", "8.43"))
                                # MEMORY OPTIMIZATION: Limit to 11 cols (actual data is ~7 cols)
                                for c in range(min_c, min(max_c + 1, 11)):
                                    widths[c] = w
                                elem.clear()
                            elif tag == f"{_SSML}mergeCell":
                                ref = elem.get("ref")
                                if ref:
                                    merges.append(ref)
                                elem.clear()
                            elif tag == f"{_SSML}row":
                                row_num = elem.get("r")
                                if row_num:
                                    row_idx = int(row_num)
                                    # MEMORY FIX: Only capture row heights in our data range
                                    if row_idx <= MAX_STYLE_ROWS:
                                        ht = elem.get("ht")
                                        if ht:
                                            heights[row_idx] = float(ht)
                                elem.clear()
                            elif tag == f"{_SSML}c":
                                # Cell element - extract style reference
                                # MEMORY FIX: Only capture style for cells in our data range
                                cell_ref = elem.get("r")
                                style_id = elem.get("s")
                                if cell_ref and style_id:
                                    row_idx, col_idx = ExcelHandler._ref_to_rowcol(cell_ref)
                                    if row_idx <= MAX_STYLE_ROWS:
                                        formats[(row_idx, col_idx)] = {"styleId": int(style_id)}
                                elem.clear()
                            elif tag in (f"{_SSML}sheetData", f"{_SSML}cols", f"{_SSML}mergeCells"):
                                elem.clear()

                    column_widths[sheet_name] = widths
                    merged_ranges[sheet_name] = merges
                    row_heights[sheet_name] = heights
                    cell_formats[sheet_name] = formats
        except Exception as e:
            print(f"Warning: Could not parse XLSX layout from ZIP: {e}")

        return column_widths, merged_ranges, row_heights, cell_formats, style_definitions

    # ── read_template ────────────────────────────────────────────────

    @staticmethod
    def read_template(
        excel_path: Path,
        target_tabs: list[str] | None = None,
    ) -> tuple[None, ExcelTemplate]:
        """
        Read an Excel template using read_only mode (streaming, constant memory).

        Also parses column widths and merged cells directly from the ZIP.
        Returns (None, template) — fill_template writes via xlsxwriter.

        Args:
            excel_path: Path to the Excel file
            target_tabs: Optional substrings to filter sheet names
                         (e.g. ["1.0", "2.0.b"])

        Returns:
            Tuple of (None, ExcelTemplate)
        """
        sheet_names: list[str] = []
        headers: dict[str, list[str]] = {}
        header_to_col: dict[str, dict[str, int]] = {}
        header_rows: dict[str, int] = {}
        structure: dict[str, list[dict[str, Any]]] = {}
        sheet_types: dict[str, str] = {}
        form_fields: dict[str, list[dict[str, Any]]] = {}
        all_existing_cells: dict[str, dict[tuple[int, int], Any]] = {}
        max_rows: dict[str, int] = {}
        max_cols: dict[str, int] = {}

        # ── Phase 1: openpyxl read_only for structure + cell values ──
        wb = openpyxl.load_workbook(
            excel_path, read_only=True, data_only=True,
            keep_vba=False, keep_links=False,
        )
        try:
            all_sheet_names = wb.sheetnames

            # Filter to target tabs
            if target_tabs:
                sheet_names = [
                    n for n in all_sheet_names
                    if any(pat.lower() in n.lower() for pat in target_tabs)
                ]
                skipped = set(all_sheet_names) - set(sheet_names)
                for s in skipped:
                    print(f"  Skipping non-target sheet '{s}'")
            else:
                sheet_names = list(all_sheet_names)

            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                sheet_headers: list[str] = []
                sheet_header_to_col: dict[str, int] = {}
                sheet_form_fields: list[dict[str, Any]] = []
                cells: dict[tuple[int, int], Any] = {}
                max_r = 0
                max_c = 0

                # Single pass: capture all cells (up to 50 rows, 10 cols)
                # OPTIMIZED: Actual data is ~36 rows × 7 cols, using 50×10 for safety buffer
                rows_by_idx: dict[int, list[Any]] = {}
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=50, max_col=10,
                                 values_only=True), 1
                ):
                    row_list = list(row) if row else []
                    rows_by_idx[row_idx] = row_list
                    for col_idx, val in enumerate(row_list, 1):
                        if val is not None:
                            cells[(row_idx, col_idx)] = val
                            max_r = max(max_r, row_idx)
                            max_c = max(max_c, col_idx)

                all_existing_cells[sheet_name] = cells
                max_rows[sheet_name] = max_r
                max_cols[sheet_name] = max_c

                # ── Detect form vs table (first 30 rows) ──
                form_like_rows = sum(
                    1
                    for r in range(1, min(31, len(rows_by_idx) + 1))
                    if rows_by_idx.get(r)
                    and len([v for v in rows_by_idx[r] if v is not None]) == 1
                    and (rows_by_idx[r][0] is not None if rows_by_idx[r] else False)
                )
                table_like_rows = sum(
                    1
                    for r in range(1, min(31, len(rows_by_idx) + 1))
                    if rows_by_idx.get(r)
                    and len([v for v in rows_by_idx[r] if v is not None]) >= 3
                )
                best_header_row = None
                best_header_count = 0
                for r in range(1, min(31, len(rows_by_idx) + 1)):
                    row = rows_by_idx.get(r, [])
                    n = len([v for v in row if v is not None])
                    if n >= 3 and n > best_header_count:
                        best_header_count = n
                        best_header_row = r

                is_form = form_like_rows > table_like_rows * 2
                sheet_types[sheet_name] = "form" if is_form else "table"

                # MEMORY FIX: Process form fields while we still have rows_by_idx
                if is_form:
                    # OPTIMIZED: Check up to 50 rows (actual data ~36 rows)
                    for row_idx in range(1, min(50, len(rows_by_idx) + 1)):
                        row = rows_by_idx.get(row_idx, [])
                        label = row[0] if len(row) > 0 else None
                        if label and isinstance(label, str) and len(label.strip()) > 0:
                            existing_values = {}
                            for col in range(1, min(9, len(row))):
                                if len(row) > col and row[col] is not None:
                                    existing_values[col + 2] = row[col]
                            sheet_form_fields.append({
                                "row": row_idx,
                                "label": label.strip(),
                                "answer_col": 2,
                                "existing": existing_values,
                            })
                    form_fields[sheet_name] = sheet_form_fields
                    sheet_headers = ["Label", "Answer", "Notes", "Reference"]
                    for i, h in enumerate(sheet_headers, 1):
                        sheet_header_to_col[h] = i
                        sheet_header_to_col[h.lower()] = i
                    header_rows[sheet_name] = 1
                else:
                    found_header_row = best_header_row or 1
                    header_row_data = rows_by_idx.get(found_header_row, [])
                    for col_idx, v in enumerate(header_row_data, 1):
                        header_name = str(v).strip() if v else f"Column_{col_idx}"
                        header_name = " ".join(header_name.split())
                        sheet_headers.append(header_name)
                        sheet_header_to_col[header_name] = col_idx
                        sheet_header_to_col[header_name.lower()] = col_idx
                        clean_name = "".join(
                            c for c in header_name.lower()
                            if c.isalnum() or c.isspace()
                        )
                        sheet_header_to_col[clean_name] = col_idx
                    header_rows[sheet_name] = found_header_row

                headers[sheet_name] = sheet_headers
                header_to_col[sheet_name] = sheet_header_to_col
                structure[sheet_name] = []

                # MEMORY FIX: Delete rows_by_idx to free memory immediately
                del rows_by_idx
                gc.collect()
        finally:
            wb.close()
        del wb
        gc.collect()

        # ── Phase 2: Parse layout from ZIP (column widths, merges, formatting) ──
        col_widths, merge_ranges, row_hts, cell_fmts, style_defs = ExcelHandler._parse_layout_from_zip(
            excel_path, sheet_names,
        )
        log_memory("after ZIP layout parse")

        return None, ExcelTemplate(
            filepath=excel_path,
            sheet_names=sheet_names,
            headers=headers,
            header_to_col=header_to_col,
            header_row=header_rows,
            structure=structure,
            sheet_type=sheet_types,
            form_fields=form_fields,
            column_widths=col_widths,
            merged_ranges=merge_ranges,
            existing_cells=all_existing_cells,
            max_row=max_rows,
            max_col=max_cols,
            row_heights=row_hts,
            cell_formats=cell_fmts,
            style_definitions=style_defs,
        )

    # ── fill_template (xlsxwriter) ───────────────────────────────────

    @staticmethod
    def _create_format_from_style(
        wb: Any,  # xlsxwriter.Workbook
        style_id: int,
        style_defs: dict[str, Any],
        confidence_override: str | None = None,
    ) -> Any:  # xlsxwriter.Format
        """
        Create an xlsxwriter format object from parsed style definitions.
        If confidence_override is provided ("low" or "medium"), override the background color.
        """
        fmt_props = {}

        cellXfs = style_defs.get("cellXfs", [])
        if style_id < len(cellXfs):
            xf = cellXfs[style_id]

            # Font
            font_id = xf.get("fontId")
            if font_id is not None and font_id < len(style_defs.get("fonts", [])):
                font = style_defs["fonts"][font_id]
                if "name" in font:
                    fmt_props["font_name"] = font["name"]
                if "size" in font:
                    fmt_props["font_size"] = font["size"]
                if font.get("bold"):
                    fmt_props["bold"] = True
                if font.get("italic"):
                    fmt_props["italic"] = True
                if "color" in font:
                    fmt_props["font_color"] = f"#{font['color']}"

            # Fill (background color) - override if confidence specified
            if confidence_override == "low":
                fmt_props["bg_color"] = "#FFCCCC"
            elif confidence_override == "medium":
                fmt_props["bg_color"] = "#FFFFCC"
            else:
                fill_id = xf.get("fillId")
                if fill_id is not None and fill_id < len(style_defs.get("fills", [])):
                    fill = style_defs["fills"][fill_id]
                    if fill.get("fgColor") and fill.get("patternType") in ("solid", "darkGray", "lightGray"):
                        fmt_props["bg_color"] = f"#{fill['fgColor']}"

            # Borders
            border_id = xf.get("borderId")
            if border_id is not None and border_id < len(style_defs.get("borders", [])):
                border = style_defs["borders"][border_id]
                border_map = {
                    "thin": 1, "medium": 2, "thick": 5, "double": 6,
                    "hair": 7, "dotted": 4, "dashed": 3, "dashDot": 9,
                    "dashDotDot": 10, "slantDashDot": 11,
                }
                for side in ["left", "right", "top", "bottom"]:
                    if side in border:
                        fmt_props[f"{side}"] = border_map.get(border[side], 1)

            # Alignment
            alignment = xf.get("alignment", {})
            h_align_map = {
                "left": "left", "center": "center", "right": "right",
                "fill": "fill", "justify": "justify", "distributed": "distributed",
            }
            v_align_map = {
                "top": "top", "center": "vcenter", "bottom": "bottom",
                "justify": "vjustify", "distributed": "vdistributed",
            }
            if "horizontal" in alignment:
                fmt_props["align"] = h_align_map.get(alignment["horizontal"], "left")
            if "vertical" in alignment:
                fmt_props["valign"] = v_align_map.get(alignment["vertical"], "top")
            if alignment.get("wrapText"):
                fmt_props["text_wrap"] = True

            # Number format
            num_fmt_id = xf.get("numFmtId")
            if num_fmt_id is not None:
                # Check custom formats
                if num_fmt_id in style_defs.get("numFmts", {}):
                    fmt_props["num_format"] = style_defs["numFmts"][num_fmt_id]
                # Built-in formats (common ones)
                elif num_fmt_id == 1:
                    fmt_props["num_format"] = "0"
                elif num_fmt_id == 2:
                    fmt_props["num_format"] = "0.00"
                elif num_fmt_id == 14:
                    fmt_props["num_format"] = "mm/dd/yyyy"

        return wb.add_format(fmt_props)

    @staticmethod
    def fill_template(
        template: ExcelTemplate,
        mappings: dict[str, list[dict[str, Any]]],
        output_path: Path,
    ) -> Path:
        """
        Write a new Excel file with xlsxwriter, reproducing the template
        layout (column widths, merged cells, row heights, fonts, colors, borders)
        and overlaying the AI-extracted data with confidence-based coloring.

        xlsxwriter writes forward-only, row-by-row, so memory usage stays
        constant regardless of sheet size.
        """
        import xlsxwriter

        print(f"\n{'='*60}")
        print("FILL_TEMPLATE (xlsxwriter)")
        print(f"{'='*60}")
        print(f"Target sheets: {template.sheet_names}")
        print(f"Mappings for: {list(mappings.keys())}")
        print(f"Sheet types: {template.sheet_type}")

        wb = xlsxwriter.Workbook(str(output_path))
        style_defs = template.style_definitions

        # Cache format objects by (styleId, confidence_override)
        format_cache: dict[tuple[int | None, str | None], Any] = {}

        def get_format(style_id: int | None, confidence: str | None = None) -> Any:
            """Get or create a format object for the given style and confidence."""
            cache_key = (style_id, confidence)
            if cache_key not in format_cache:
                if style_id is not None:
                    format_cache[cache_key] = ExcelHandler._create_format_from_style(
                        wb, style_id, style_defs, confidence
                    )
                else:
                    # Default format
                    props = {}
                    if confidence == "low":
                        props["bg_color"] = "#FFCCCC"
                    elif confidence == "medium":
                        props["bg_color"] = "#FFFFCC"
                    format_cache[cache_key] = wb.add_format(props)
            return format_cache[cache_key]

        def normalize_confidence(c: Any) -> str:
            if c in ("h", "high"):
                return "high"
            if c in ("m", "medium", "med"):
                return "medium"
            if c in ("l", "low"):
                return "low"
            return "high"

        for sheet_name in template.sheet_names:
            ws = wb.add_worksheet(sheet_name)
            sheet_type = template.sheet_type.get(sheet_name, "table")
            h_row = template.header_row.get(sheet_name, 1)
            header_map = template.header_to_col.get(sheet_name, {})
            existing = template.existing_cells.get(sheet_name, {})
            max_r = template.max_row.get(sheet_name, 0)
            max_c = template.max_col.get(sheet_name, 0)
            cell_fmts = template.cell_formats.get(sheet_name, {})

            # ── Set column widths ──
            for col_1, width in template.column_widths.get(sheet_name, {}).items():
                ws.set_column(col_1 - 1, col_1 - 1, width)

            # ── Set row heights ──
            for row_1, height in template.row_heights.get(sheet_name, {}).items():
                ws.set_row(row_1 - 1, height)

            # ── Step A: Build AI cell overlay FIRST ──
            # (must happen before merge writing so AI values can override)
            ai_rows: list[dict[str, Any]] = []
            matched = ExcelHandler._match_sheet_name(sheet_name, list(mappings.keys()))
            if matched is None:
                for ai_name in mappings:
                    if ExcelHandler._match_sheet_name(ai_name, [sheet_name]):
                        matched = ai_name
                        break
            if matched is not None:
                ai_rows = mappings[matched]
                print(f"\n  Sheet '{sheet_name}': {len(ai_rows)} AI rows from key '{matched}'")

            ai_cells: dict[tuple[int, int], tuple[Any, str]] = {}
            for row_idx_in, row_update in enumerate(ai_rows):
                row_idx = row_update.get("_row")
                if row_idx is None:
                    continue
                confidence_map = row_update.get("_confidence", row_update.get("_c", {}))
                row_conf = row_update.get("_row_confidence", "high")

                if sheet_type == "form":
                    answer = row_update.get("Answer") or row_update.get("answer")
                    if answer:
                        ai_cells[(row_idx, 2)] = (answer, normalize_confidence(row_conf))
                    for col_name, value in row_update.items():
                        if col_name.startswith("_") or col_name.lower() == "answer":
                            continue
                        if value and isinstance(value, str):
                            cidx = header_map.get(col_name) or header_map.get(col_name.lower())
                            if cidx:
                                ai_cells[(row_idx, cidx)] = (value, normalize_confidence(row_conf))
                else:
                    if row_idx <= h_row:
                        row_idx = h_row + 1 + row_idx_in
                    for col_name, value in row_update.items():
                        if col_name.startswith("_"):
                            continue
                        cell_value = value
                        raw_conf = confidence_map.get(col_name, "high")
                        cell_conf = normalize_confidence(raw_conf)
                        if isinstance(value, dict):
                            cell_value = value.get("value", value.get("v"))
                            raw_conf = value.get("confidence", value.get("c", raw_conf))
                            cell_conf = normalize_confidence(raw_conf)
                        if cell_value is None or cell_value == "":
                            continue
                        col_idx = ExcelHandler._resolve_col(col_name, header_map)
                        if col_idx:
                            ai_cells[(row_idx, col_idx)] = (cell_value, cell_conf)
                            max_r = max(max_r, row_idx)
                            max_c = max(max_c, col_idx)

            # ── Step B: Write merged ranges ──
            # AI data overrides existing values for any merge whose top-left
            # cell has an AI entry (e.g. form answer areas in the Mgmt Review).
            merged_set: set[tuple[int, int]] = set()
            cells_written_total = 0
            for merge_ref in template.merged_ranges.get(sheet_name, []):
                fr, fc, lr, lc = ExcelHandler._parse_range(merge_ref)
                for mr in range(fr, lr + 1):
                    for mc in range(fc, lc + 1):
                        merged_set.add((mr, mc))

                # Check if AI has data for ANY cell in this merge range.
                # If so, use the first AI hit as the merge value.
                ai_val = None
                ai_conf = "high"
                for mr in range(fr, lr + 1):
                    for mc in range(fc, lc + 1):
                        if (mr, mc) in ai_cells:
                            ai_val, ai_conf = ai_cells[(mr, mc)]
                            break
                    if ai_val is not None:
                        break

                # Get original cell style from template
                orig_style_id = cell_fmts.get((fr, fc), {}).get("styleId")

                if ai_val is not None:
                    val = ai_val
                    # AI data: apply confidence coloring (overrides original bg)
                    conf_override = ai_conf if ai_conf != "high" else None
                    fmt = get_format(orig_style_id, conf_override)
                    cells_written_total += 1
                else:
                    # No AI data: use original template value and style
                    val = existing.get((fr, fc), "")
                    fmt = get_format(orig_style_id)

                if fr == lr and fc == lc:
                    ws.write(fr - 1, fc - 1, val, fmt)
                else:
                    ws.merge_range(fr - 1, fc - 1, lr - 1, lc - 1, val, fmt)

            # ── Step C: Write non-merged cells ──
            for r in range(1, max_r + 1):
                for c in range(1, max_c + 1):
                    if (r, c) in merged_set:
                        continue  # Already handled by merge_range above

                    # Get original cell style from template
                    orig_style_id = cell_fmts.get((r, c), {}).get("styleId")

                    # AI data takes priority over existing template data
                    if (r, c) in ai_cells:
                        val, conf = ai_cells[(r, c)]
                        # AI data: apply confidence coloring (overrides original bg)
                        conf_override = conf if conf != "high" else None
                        fmt = get_format(orig_style_id, conf_override)
                        ws.write(r - 1, c - 1, val, fmt)
                        cells_written_total += 1
                    elif (r, c) in existing:
                        # No AI data: use original template value and style
                        val = existing[(r, c)]
                        fmt = get_format(orig_style_id)
                        ws.write(r - 1, c - 1, val, fmt)

            print(f"  Sheet '{sheet_name}': wrote {cells_written_total} AI cells")

            # MEMORY FIX: Aggressive GC after each sheet
            gc.collect()

        wb.close()
        log_memory("after xlsxwriter close")
        return output_path

    @staticmethod
    def _resolve_col(col_name: str, header_map: dict[str, int]) -> int | None:
        """Resolve a column name to a 1-based column index using multiple strategies."""
        # Strategy 1: Exact match
        idx = header_map.get(col_name)
        if idx is not None:
            return idx

        # Strategy 2: Case-insensitive
        idx = header_map.get(col_name.lower().strip())
        if idx is not None:
            return idx

        # Strategy 3: Partial / substring match
        col_clean = col_name.lower().strip()
        for hk, hi in header_map.items():
            if isinstance(hk, str):
                hk_clean = hk.lower().strip()
                if col_clean in hk_clean or hk_clean in col_clean:
                    return hi

        # Strategy 4: Word overlap (≥2 common words)
        col_words = set(col_name.lower().split())
        best_score, best_idx = 0, None
        for hk, hi in header_map.items():
            if isinstance(hk, str):
                common = col_words & set(hk.lower().split())
                if len(common) > best_score and len(common) >= 2:
                    best_score = len(common)
                    best_idx = hi
        return best_idx


class SOC1Agent:
    """
    AI-powered agent for processing SOC1 Type II reports.

    Uses Google Gemini (free tier available) for intelligent content extraction.
    """

    def __init__(self, api_key: str | None = None):
        """
        Initialize the SOC1 Agent.

        Args:
            api_key: Google API key. If not provided, uses GOOGLE_API_KEY env var.
                     Get a free key at: https://aistudio.google.com/apikey
        """
        self.api_key = api_key or os.environ.get("GOOGLE_API_KEY")
        if not self.api_key:
            raise ValueError(
                "Google API key required. Set GOOGLE_API_KEY environment variable "
                "or pass api_key parameter.\n"
                "Get a free key at: https://aistudio.google.com/apikey"
            )
        # Initialize the genai client with API key
        self.client = genai.Client(api_key=self.api_key)
        # Use Gemini 2.5 Flash for free tier (latest, fast and capable)
        self.model = "gemini-2.5-flash"

    def _generate(self,
        prompt: str,
        max_tokens: int = 8192,
        retries: int = 5,  # PERFORMANCE: Increased from 3 to 5 for better reliability
        temperature: float | None = None,  # PERFORMANCE: Allow caller to specify temperature
        expect_json: bool = False) -> str:
        """Generate a response from Gemini with retry logic."""
        import time

        last_error = None
        for attempt in range(retries):
            try:
                # PERFORMANCE: Higher temperature (0.3) for better reasoning
                # Lower temp (0.1) was too conservative, missed valid inferences
                if temperature is None:
                    temperature = 0.3 if max_tokens > 10000 else 0.2

                config_kwargs = dict(
                    max_output_tokens=max_tokens,
                    temperature=temperature,
                )
                if expect_json:
                    config_kwargs["response_mime_type"] = "application/json"

                response = self.client.models.generate_content(
                    model=self.model,
                    contents=prompt,
                    config=types.GenerateContentConfig(**config_kwargs),
                )
                
                # Check if response has text
                if getattr(response, "text", None):
                    return response.text

                # Fallback: pull text from candidates/parts when response.text is empty
                if getattr(response, "candidates", None):
                    cand0 = response.candidates[0]
                    content = getattr(cand0, "content", None)
                    parts = getattr(content, "parts", None) if content else None
                    if parts:
                        for p in parts:
                            t = getattr(p, "text", None)
                            if t:
                                return t

                raise ValueError("Empty response from Gemini")
                
            except Exception as e:
                last_error = e
                print(f"Attempt {attempt + 1}/{retries} failed: {str(e)}")
                if attempt < retries - 1:
                    wait_time = (attempt + 1) * 2  # Exponential backoff
                    print(f"Waiting {wait_time}s before retry...")
                    time.sleep(wait_time)
        
        raise RuntimeError(f"Failed after {retries} attempts. Last error: {last_error}")

    def _parse_json_response(self, response_text: str) -> dict[str, Any]:
        """Parse JSON from AI response, handling markdown code blocks and incomplete JSON."""
        original_text = response_text
        
        # Step 1: Remove markdown code block markers (handle both complete and incomplete blocks)
        # First try to match complete code blocks
        json_match = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", response_text)
        if json_match:
            response_text = json_match.group(1).strip()
        else:
            # Handle incomplete code blocks (no closing ```)
            if response_text.strip().startswith("```"):
                # Remove opening ``` and optional json tag
                response_text = re.sub(r"^```(?:json)?\s*", "", response_text.strip())
            # Also remove any trailing ``` if present
            response_text = re.sub(r"\s*```\s*$", "", response_text)

        # Step 2: Try direct parsing
        try:
            return json.loads(response_text)
        except json.JSONDecodeError:
            pass

        # Step 3: Find the JSON object boundaries
        json_start = response_text.find("{")
        if json_start < 0:
            print(f"No JSON object found in response: {original_text[:500]}")
            raise ValueError(f"No JSON object found in AI response. Response: {original_text[:200]}...")

        # Step 4: Try to extract valid JSON
        json_candidate = response_text[json_start:]
        
        # Try parsing as-is first
        try:
            return json.loads(json_candidate)
        except json.JSONDecodeError:
            pass

        # Step 5: Try to repair truncated JSON
        # Find the last valid closing brace
        json_candidate = self._repair_truncated_json(json_candidate)
        
        try:
            result = json.loads(json_candidate)
            print(f"Successfully parsed JSON after repair (found {len(result)} keys)")
            return result
        except json.JSONDecodeError as e:
            # Log detailed error for debugging
            print(f"Failed to parse JSON response after repair attempts.")
            print(f"JSON error: {e.msg} at position {e.pos}")
            print(f"Context around error: ...{json_candidate[max(0,e.pos-50):e.pos+50]}...")
            print(f"Original text length: {len(original_text)}")
            print(f"Repaired text length: {len(json_candidate)}")
            
            # Last resort: try to extract just the first complete sheet
            try:
                # Find the first complete array in the JSON
                first_array_end = json_candidate.find(']')
                if first_array_end > 0:
                    partial = json_candidate[:first_array_end + 1] + '}'
                    result = json.loads(partial)
                    print(f"Recovered partial JSON with {len(result)} keys")
                    return result
            except:
                pass
            
            raise ValueError(
                f"Could not parse AI response as JSON. The response may be truncated. "
                f"Response preview: {original_text[:300]}..."
            )

    def _repair_truncated_json(self, json_str: str) -> str:
        """Attempt to repair truncated JSON by closing open brackets and braces."""
        # Strategy: Find the last complete row entry and truncate there
        
        # First, try to find the last complete object in an array
        # Look for pattern: }, { which indicates complete objects in array
        last_complete = json_str.rfind('},')
        if last_complete > 0:
            # Check if cutting here gives us valid-ish JSON
            test_str = json_str[:last_complete + 1]  # Include the }
            open_braces = test_str.count('{') - test_str.count('}')
            open_brackets = test_str.count('[') - test_str.count(']')
            
            # If we have more opens than closes, this might be a good cut point
            if open_braces >= 0 and open_brackets >= 0:
                json_str = test_str
        
        # Remove any trailing incomplete content
        # Pattern 1: Incomplete key-value with nested object: "key": {incomplete
        json_str = re.sub(r',\s*"[^"]*"\s*:\s*\{[^{}]*$', '', json_str)
        # Pattern 2: Incomplete key-value with string: "key": "incomplete
        json_str = re.sub(r',\s*"[^"]*"\s*:\s*"[^"]*$', '', json_str)
        # Pattern 3: Incomplete key-value with object having confidence: "key": {"value": "x", "confidence
        json_str = re.sub(r',\s*"[^"]*"\s*:\s*\{[^{}]*$', '', json_str)
        # Pattern 4: Just an incomplete key: , "key
        json_str = re.sub(r',\s*"[^"]*$', '', json_str)
        # Pattern 5: Incomplete array element
        json_str = re.sub(r',\s*\[[^\[\]]*$', '', json_str)
        json_str = re.sub(r',\s*\{[^{}]*$', '', json_str)
        
        # Remove trailing comma if present
        json_str = re.sub(r',\s*$', '', json_str)
        
        # Count and balance brackets/braces
        open_braces = json_str.count('{') - json_str.count('}')
        open_brackets = json_str.count('[') - json_str.count(']')
        
        print(f"Repair: {open_brackets} unclosed brackets, {open_braces} unclosed braces")
        
        # Close any remaining open structures (close brackets before braces)
        for _ in range(max(0, open_brackets)):
            json_str += ']'
        for _ in range(max(0, open_braces)):
            json_str += '}'
        
        return json_str

    def _create_extraction_prompt(
        self,
        pdf_content: ExtractedPDFContent,
        template: ExcelTemplate,
        target_sheets: list[str] | None = None,
    ) -> str:
        """Create a prompt for the AI to extract and map data.

        Args:
            pdf_content: Extracted PDF content
            template: Excel template structure
            target_sheets: Optional list of sheet name substrings to filter which sheets to process
        """

        # Find sheets that match management review and CUEC patterns
        # Filter by target_sheets if provided
        management_sheet = None
        cuec_sheet = None
        deviations_sheet = None

        for sheet_name in template.sheet_names:
            # Skip if target_sheets is specified and this sheet doesn't match
            if target_sheets and not any(target in sheet_name for target in target_sheets):
                continue

            sheet_lower = sheet_name.lower()
            if "management review" in sheet_lower and "1.0" in sheet_name:
                management_sheet = sheet_name
            if "user entity" in sheet_lower or "cuec" in sheet_lower or "comp user" in sheet_lower:
                cuec_sheet = sheet_name
            if "deviation" in sheet_lower:
                deviations_sheet = sheet_name

        prompt_parts = []
        prompt_parts.append("You are an expert at extracting data from SOC1 Type II audit reports to fill Excel management review templates.")
        prompt_parts.append("\n\n## EXTRACTION STRATEGY:")
        prompt_parts.append("1. Read the entire document to understand structure and sections")
        prompt_parts.append("2. For each data point, cross-reference across pages if needed")
        prompt_parts.append("3. Synthesize information that spans multiple sections")
        prompt_parts.append("4. Use 'medium' confidence when combining data from different pages")
        prompt_parts.append("5. Use 'low' confidence when inferring from context")

        prompt_parts.append("\n\n## PDF REPORT CONTENT:\n")
        # PERFORMANCE: Expanded from 80K to 200K chars for 2.5x more context
        prompt_parts.append(pdf_content.full_text[:200000])

        if pdf_content.tables:
            prompt_parts.append("\n\n## EXTRACTED TABLES FROM PDF:\n")
            # PERFORMANCE: Show all tables (was limited to 15), more rows per table (25 vs 10)
            for i, table in enumerate(pdf_content.tables[:40], 1):
                prompt_parts.append(f"\nTable {i}:")
                for row in table[:25]:
                    prompt_parts.append(f"  {row}")
        
        prompt_parts.append("\n\n## EXCEL SHEETS TO FILL:\n")
        
        # Handle Management Review sheet (form-style)
        if management_sheet and template.sheet_type.get(management_sheet) == "form":
            form_fields = template.form_fields.get(management_sheet, [])
            prompt_parts.append(f"\n### Sheet: {management_sheet} (FORM-STYLE)")
            prompt_parts.append("This is a questionnaire. For each question, provide the answer from the PDF.")
            prompt_parts.append("Fields to fill:")
            # PERFORMANCE: Show ALL fields (was limited to 30)
            for field in form_fields:
                label = field['label'][:100]  # Show more of label (was 80 chars)
                prompt_parts.append(f"  - Row {field['row']}: \"{label}\"")
        
        # Handle CUEC sheet (table-style)
        if cuec_sheet and template.sheet_type.get(cuec_sheet) == "table":
            cuec_headers = template.headers.get(cuec_sheet, [])
            prompt_parts.append(f"\n### Sheet: {cuec_sheet} (TABLE)")
            prompt_parts.append(f"Header row: {template.header_row.get(cuec_sheet)}")
            prompt_parts.append("Column headers (USE THESE EXACT NAMES):")
            # PERFORMANCE: Show ALL headers (was limited to 10)
            for i, h in enumerate(cuec_headers, 1):
                prompt_parts.append(f"  {i}. \"{h}\"")
            prompt_parts.append("\nLook for 'Complementary User Entity Controls' section in the PDF.")
            prompt_parts.append("Extract ALL CUECs - they are often in tables or numbered lists near the end of the document.")
            prompt_parts.append("CUECs may also be called 'User Entity Responsibilities' or 'Subservice Organization Controls'.")

        # Handle Deviations sheet
        if deviations_sheet and template.sheet_type.get(deviations_sheet) == "table":
            dev_headers = template.headers.get(deviations_sheet, [])
            prompt_parts.append(f"\n### Sheet: {deviations_sheet} (TABLE)")
            prompt_parts.append(f"Header row: {template.header_row.get(deviations_sheet)}")
            prompt_parts.append("Column headers:")
            # PERFORMANCE: Show ALL headers (was limited to 10)
            for i, h in enumerate(dev_headers, 1):
                prompt_parts.append(f"  {i}. \"{h}\"")
        
        # JSON format instructions
        prompt_parts.append("\n\n## RETURN FORMAT:")
        prompt_parts.append("Return ONLY valid JSON. No markdown code blocks, no commentary.")
        prompt_parts.append("{")
        
        if management_sheet:
            prompt_parts.append(f'  "{management_sheet}": [')
            prompt_parts.append('    {"_row": 4, "Answer": "Okta, Inc."},  // Row 4 answer')
            prompt_parts.append('    {"_row": 5, "Answer": "SOC1 Type II Report"},')
            prompt_parts.append("    ...")
            prompt_parts.append("  ],")
        
        if cuec_sheet:
            cuec_h = template.headers.get(cuec_sheet, ["No.", "Description"])
            h1 = cuec_h[0] if len(cuec_h) > 0 else "No."
            h2 = cuec_h[2] if len(cuec_h) > 2 else "Description"
            h3 = cuec_h[3] if len(cuec_h) > 3 else "Control Objective"
            start_row = template.header_row.get(cuec_sheet, 1) + 1
            prompt_parts.append(f'  "{cuec_sheet}": [')
            prompt_parts.append(f'    {{"_row": {start_row}, "{h1}": "1", "{h2}": "User entities are responsible for...", "{h3}": "CO 2 - Logical access"}},')
            prompt_parts.append(f'    {{"_row": {start_row + 1}, "{h1}": "2", "{h2}": "Another CUEC...", "{h3}": "CO 2 - Logical access"}}')
            prompt_parts.append("  ]")
        
        prompt_parts.append("}")
        
        prompt_parts.append("\n\n## IMPORTANT RULES:")
        prompt_parts.append("1. Use EXACT column names from the headers listed above")
        prompt_parts.append("2. For table sheets, _row is the row number (starts after header row)")
        prompt_parts.append("3. For form sheets, _row matches the row of the question label")
        prompt_parts.append("4. Extract ALL CUECs from the 'Complementary User Entity Controls' section")
        prompt_parts.append("5. Keep values concise but complete")
        prompt_parts.append("6. Return valid JSON only - no markdown, no extra text")
        
        return "\n".join(prompt_parts)

    def extract_and_map(
        self,
        pdf_content: ExtractedPDFContent,
        template: ExcelTemplate,
        target_sheets: list[str] | None = None,
    ) -> dict[str, list[dict[str, Any]]]:
        """
        Use AI to extract data from PDF and map to Excel template.

        Args:
            pdf_content: Extracted PDF content
            template: Excel template structure
            target_sheets: Optional list of sheet name substrings to filter which sheets to process

        Returns:
            Dict mapping sheet names to lists of row updates
        """
        try:
            prompt = self._create_extraction_prompt(pdf_content, template, target_sheets)
            print(f"\n{'='*60}")
            print("AI EXTRACTION")
            print(f"{'='*60}")
            print(f"Prompt length: {len(prompt)} chars")
            print(f"Template sheets: {template.sheet_names}")
            print(f"Sheet types: {template.sheet_type}")

            # PERFORMANCE: Increased max_tokens to 100K (was 65K) for larger templates
            # Using temperature=0.3 (default) for better reasoning
            response_text = self._generate(prompt, max_tokens=100000, expect_json=True)
            
            print(f"\nAI Response length: {len(response_text)} chars")
            print(f"Response preview: {response_text[:500]}...")
            
            result = self._parse_json_response(response_text)
            
            print(f"\nParsed result keys: {list(result.keys())}")
            for sheet, rows in result.items():
                print(f"  {sheet}: {len(rows)} rows")
                if rows:
                    print(f"    First row keys: {list(rows[0].keys())}")
            
            return result
        except Exception as e:
            print(f"\nMain extraction failed: {e}")
            import traceback
            traceback.print_exc()
            print("Trying simplified extraction...")
            return self._extract_simplified(pdf_content, template, target_sheets)

    def _extract_simplified(
        self,
        pdf_content: ExtractedPDFContent,
        template: ExcelTemplate,
        target_sheets: list[str] | None = None,
    ) -> dict[str, list[dict[str, Any]]]:
        """Simplified extraction as fallback - extracts one sheet at a time.

        Args:
            pdf_content: Extracted PDF content
            template: Excel template structure
            target_sheets: Optional list of sheet name substrings to filter which sheets to process
        """
        result = {}

        for sheet_name in template.sheet_names:
            # Skip if target_sheets is specified and this sheet doesn't match
            if target_sheets and not any(target in sheet_name for target in target_sheets):
                continue

            sheet_lower = sheet_name.lower()

            # Only process relevant sheets
            if not any(x in sheet_lower for x in ["management review", "user entity", "cuec", "comp user"]):
                continue
            
            headers = template.headers.get(sheet_name, [])
            header_row = template.header_row.get(sheet_name, 1)
            sheet_type = template.sheet_type.get(sheet_name, "table")
            
            print(f"\n--- Simplified extraction for '{sheet_name}' ---")
            print(f"  Sheet type: {sheet_type}, Headers: {headers[:5]}")
            
            if "user entity" in sheet_lower or "cuec" in sheet_lower or "comp user" in sheet_lower:
                # Special handling for CUEC sheet - look specifically for the CUEC section
                # Look for CUEC section in the text
                cuec_section = ""
                full_text_lower = pdf_content.full_text.lower()
                cuec_start = full_text_lower.find("complementary user entity control")
                if cuec_start == -1:
                    cuec_start = full_text_lower.find("cuec")

                if cuec_start != -1:
                    # PERFORMANCE: Extract 60K chars around CUEC section (was 20K)
                    start = max(0, cuec_start - 10000)
                    end = min(len(pdf_content.full_text), cuec_start + 50000)
                    cuec_section = pdf_content.full_text[start:end]
                else:
                    # PERFORMANCE: Use last 60K chars (was 20K) - CUECs often at end
                    cuec_section = pdf_content.full_text[-60000:]
                
                prompt = f"""Extract ALL Complementary User Entity Controls (CUECs) from this SOC1 Type II report.

CUECs describe responsibilities that the user organization must perform. They may be called:
- "Complementary User Entity Controls"
- "CUECs"
- "User Entity Responsibilities"
- "Complementary Controls"
- "User Organization Controls"

Common patterns:
- Usually in a dedicated section near the end of the report
- Often in a table or numbered list format
- Each CUEC has: a number/ID, description, and related control objective
- There are typically 5-20 CUECs in a report

PDF Content (CUEC section + context):
{cuec_section}

Excel columns to fill (USE THESE EXACT NAMES):
{chr(10).join(f'  - "{h}"' for h in headers)}

The header row is {header_row}, so data starts at row {header_row + 1}.

IMPORTANT: Extract ALL CUECs you find. Read carefully through tables and lists.

Return JSON array with one object per CUEC found:
[
  {{"_row": {header_row + 1}, "No.": "1", "{headers[2] if len(headers) > 2 else 'Description'}": "User entities are responsible for...", "{headers[3] if len(headers) > 3 else 'Control Objective'}": "CO 2 - Logical access"}},
  {{"_row": {header_row + 2}, "No.": "2", "{headers[2] if len(headers) > 2 else 'Description'}": "Another control...", "{headers[3] if len(headers) > 3 else 'Control Objective'}": "CO 2 - Logical access"}}
]

Return ONLY valid JSON array. No markdown, no commentary."""

            elif "management review" in sheet_lower:
                # Management review form
                form_fields = template.form_fields.get(sheet_name, [])
                # PERFORMANCE: Show ALL fields (was limited to 25), more chars per field (100 vs 60)
                fields_text = "\n".join(f"  Row {f['row']}: {f['label'][:100]}" for f in form_fields)

                # PERFORMANCE: Use more PDF context (100K vs 50K)
                prompt = f"""Extract answers for this SOC1 Management Review questionnaire.

Tips for finding answers:
- Service organization name: Usually on cover page or Section 1
- Report period: Cover page or executive summary
- Report type: "Type I" or "Type II" on cover page
- Controls tested: Count from controls section (Section 3-4)
- Exceptions: Look for "exceptions", "deviations", or "findings" sections
- Opinion: Look for auditor's opinion section (often Section 2)

PDF Content:
{pdf_content.full_text[:100000]}

Questions to answer (row number: question):
{fields_text}

Return JSON array with answers. Leave Answer empty if not found in PDF:
[
  {{"_row": 4, "Answer": "Okta, Inc."}},
  {{"_row": 5, "Answer": "SOC1 Type II Report"}},
  ...
]

Return ONLY valid JSON array. No markdown."""
            else:
                # Generic extraction
                # PERFORMANCE: Show ALL headers (was 10), use more context (150K vs 80K)
                prompt = f"""Extract data from this SOC1 report for sheet "{sheet_name}".

Columns: {', '.join(headers)}

PDF Content:
{pdf_content.full_text[:150000]}

Return JSON array with one object per row:
[{{"_row": {header_row + 1}, "{headers[0]}": "value"}}]

Return ONLY JSON."""

            try:
                # PERFORMANCE: Use higher max_tokens (100K vs 65K) and better temperature (0.4 for fallback)
                response = self._generate(prompt, max_tokens=100000, temperature=0.4, expect_json=True)
                print(f"  Response length: {len(response)} chars")
                print(f"  Response preview: {response[:300]}...")
                
                rows = self._parse_json_response(response)
                
                # Handle both array and object responses
                if isinstance(rows, list):
                    result[sheet_name] = rows
                    print(f"  Extracted {len(rows)} rows")
                elif isinstance(rows, dict):
                    if sheet_name in rows:
                        result[sheet_name] = rows[sheet_name]
                    else:
                        result[sheet_name] = [rows] if "_row" in rows else []
            except Exception as e:
                print(f"  Failed: {e}")
                import traceback
                traceback.print_exc()
                result[sheet_name] = []
        
        return result

    def analyze_for_gaps(
        self,
        pdf_content: ExtractedPDFContent,
        filled_data: dict[str, list[dict[str, Any]]],
    ) -> dict[str, Any]:
        """
        Analyze the filled data for gaps or issues that need attention.

        Args:
            pdf_content: The original PDF content
            filled_data: The data that was filled into the template

        Returns:
            Dict containing analysis results and recommendations
        """
        prompt = f"""Analyze the following SOC1 Type II report data extraction for completeness and accuracy.

## Extracted Data (filled into management review template):
{json.dumps(filled_data, indent=2, default=str)[:50000]}

## Original Report Summary:
{pdf_content.full_text[:15000]}

## Analysis Required:
1. Identify any controls mentioned in the report that may not have been captured
2. Flag any exceptions or findings that need management attention
3. Note any missing information that should be followed up on
4. Provide a summary of the overall control environment
5. Count cells marked with low/medium confidence that need review
6. Analyze the Complementary User Entity Controls (CUECs) extraction

Return a JSON object:
{{
    "total_controls_identified": <number>,
    "controls_with_exceptions": <number>,
    "total_cuecs_identified": <number>,
    "cells_needing_review": {{
        "low_confidence": <number of cells with low confidence>,
        "medium_confidence": <number of cells with medium confidence>
    }},
    "missing_information": ["list of missing items"],
    "key_findings": ["list of key findings"],
    "cuec_findings": ["list of findings related to CUECs"],
    "recommendations": ["list of recommendations"],
    "summary": "brief summary of the SOC1 report"
}}

Return ONLY the JSON object."""

        # PERFORMANCE: Conservative temperature (0.2) for analysis, higher max_tokens (8K vs 4K)
        response_text = self._generate(prompt, max_tokens=8192, temperature=0.2, expect_json=True)

        try:
            return self._parse_json_response(response_text)
        except (json.JSONDecodeError, ValueError):
            return {"error": "Could not parse analysis", "raw": response_text[:1000]}


async def process_soc1_documents(
    type_ii_path: Path,
    management_review_path: Path,
    output_dir: Path | None = None,
    api_key: str | None = None,
) -> dict[str, Any]:
    """
    Main processing function for SOC1 Type II documents.

    Args:
        type_ii_path: Path to the Type II report PDF
        management_review_path: Path to the management review Excel template
        output_dir: Directory to save output files (defaults to same as input)
        api_key: Optional Google API key

    Returns:
        Dict containing:
            - output_path: Path to the filled Excel file
            - analysis: Analysis results and recommendations
            - status: Processing status
    """
    if output_dir is None:
        output_dir = management_review_path.parent

    output_dir.mkdir(parents=True, exist_ok=True)

    log_memory("start of process_soc1_documents")

    # Step 1: Extract PDF content
    print(f"Extracting content from PDF: {type_ii_path}")
    pdf_content = PDFExtractor.extract(type_ii_path)
    print(f"  - Extracted {len(pdf_content.pages)} pages")
    print(f"  - Found {len(pdf_content.tables)} tables")
    log_memory("after PDF extraction")

    # Step 2: Read Excel template (read_only + ZIP parse, constant memory)
    # Read ALL sheets, but will only process target tabs with AI
    print(f"Reading Excel template: {management_review_path}")
    _, template = ExcelHandler.read_template(
        management_review_path,
        target_tabs=None,  # Read all sheets
    )
    target_sheets = ["1.0", "2.0.b"]  # Only process these with AI
    print(f"  - Total sheets: {template.sheet_names}")
    print(f"  - Target sheets for AI processing: {[s for s in template.sheet_names if any(t in s for t in target_sheets)]}")
    for sheet, headers in template.headers.items():
        print(f"  - {sheet}: {len(headers)} columns")
    log_memory("after Excel template load")

    # Step 3: Initialize AI agent and process
    print("Initializing Google Gemini AI agent...")
    agent = SOC1Agent(api_key=api_key)

    print("Extracting and mapping content using AI...")
    # Only process target sheets with AI, leave others unchanged
    mappings = agent.extract_and_map(pdf_content, template, target_sheets)
    log_memory("after AI extraction")

    # Step 4: Fill the template via xlsxwriter (streaming, constant memory)
    output_filename = f"filled_{management_review_path.name}"
    output_path = output_dir / output_filename

    print("Filling Excel template...")
    ExcelHandler.fill_template(template, mappings, output_path)
    print(f"  - Saved to: {output_path}")
    log_memory("after xlsxwriter save")

    gc.collect()
    log_memory("after gc")

    # Step 5: Analyze for gaps (only needs a subset of PDF text)
    # Trim pdf_content to reduce memory before gap analysis
    pdf_metadata = pdf_content.metadata
    template_sheets = template.sheet_names
    pdf_text_for_analysis = pdf_content.full_text[:10000]

    # Free the bulk of PDF content
    pdf_content.full_text = ""
    pdf_content.pages.clear()
    pdf_content.tables.clear()
    del template
    gc.collect()
    log_memory("after freeing PDF content and template")

    print("Analyzing extraction for completeness...")
    # Build a lightweight PDF content object for analysis
    analysis_pdf = ExtractedPDFContent(
        full_text=pdf_text_for_analysis,
        pages=[],
        tables=[],
        metadata=pdf_metadata,
    )
    analysis = agent.analyze_for_gaps(analysis_pdf, mappings)
    del analysis_pdf, pdf_text_for_analysis
    gc.collect()
    log_memory("after gap analysis")

    return {
        "output_path": str(output_path),
        "analysis": analysis,
        "status": "completed",
        "pdf_metadata": pdf_metadata,
        "template_sheets": template_sheets,
    }


# Convenience function for synchronous usage
def process_soc1_sync(
    type_ii_path: str | Path,
    management_review_path: str | Path,
    output_dir: str | Path | None = None,
    api_key: str | None = None,
) -> dict[str, Any]:
    """
    Synchronous wrapper for process_soc1_documents.

    Args:
        type_ii_path: Path to the Type II report PDF
        management_review_path: Path to the management review Excel template
        output_dir: Directory to save output files
        api_key: Optional Google API key

    Returns:
        Dict containing processing results
    """
    import asyncio

    type_ii_path = Path(type_ii_path)
    management_review_path = Path(management_review_path)
    output_dir = Path(output_dir) if output_dir else None

    return asyncio.run(
        process_soc1_documents(
            type_ii_path,
            management_review_path,
            output_dir,
            api_key,
        )
    )
